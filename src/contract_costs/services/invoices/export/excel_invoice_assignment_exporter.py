from enum import Enum
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from contract_costs.repository.contract_repository import ContractRepository
from contract_costs.repository.cost_node_repository import CostNodeRepository
from contract_costs.repository.cost_type_repository import CostTypeRepository
from contract_costs.services.invoices.commands.invoice_command import InvoiceCommand
from contract_costs.services.invoices.dto.export.assignment_export_bundle import InvoiceAssignmentExportBundle
from contract_costs.services.invoices.export.invoice_assignment_exporter import InvoiceAssignmentExporter

import contract_costs.config as cfg

class ExcelInvoiceAssignmentExporter(InvoiceAssignmentExporter):
    def __init__(self,
                 contract_repository: ContractRepository,
                 cost_node_repository: CostNodeRepository,
                 cost_type_repository:CostTypeRepository):
        self._contract_repository = contract_repository
        self._cost_node_repository = cost_node_repository
        self._cost_type_repository = cost_type_repository

    def export(
            self,
            bundle: InvoiceAssignmentExportBundle,
            output_path: Path
    ) -> None:
        wb = Workbook()

        invoices_ws = self._write_invoices(wb, bundle.invoices)
        lines_ws = self._write_invoice_lines(wb, bundle.invoice_lines)
        buyers_ws = self._write_buyers(wb, bundle.buyers)
        sellers_ws = self._write_sellers(wb, bundle.sellers)
        contracts_ws = self._write_contracts(wb, bundle.contracts)
        cost_nodes_ws = self._write_cost_nodes(wb, bundle.cost_nodes)
        cost_types_ws = self._write_cost_types(wb, bundle.cost_types)
        tax_treatment_ws = self._write_dictionary(wb,bundle.amount_types,cfg.DICTS_TAX_TREATMENTS)
        payment_method_ws = self._write_dictionary(wb, bundle.payment_methods, cfg.DICTS_PAYMENT_METHODS)
        payment_status_ws = self._write_dictionary(wb, bundle.payment_status, cfg.DICTS_PAYMENT_STATUS)
        units_ws = self._write_dictionary(wb, bundle.units, cfg.DICTS_UNITS)
        vat_rates_ws = self._write_dictionary(wb, bundle.vat_rates, cfg.DICTS_VAT_RATES)
        actions_ws = self._write_dictionary(wb, bundle.actions, cfg.DICTS_ACTIONS)


        self._apply_dropdowns(
            invoices_ws=invoices_ws,
            lines_ws=lines_ws,
            buyers_ws=buyers_ws,
            sellers_ws=sellers_ws,
            contracts_ws=contracts_ws,
            cost_nodes_ws=cost_nodes_ws,
            cost_types_ws=cost_types_ws,
            tax_treatments_ws=tax_treatment_ws,
            payment_method_ws=payment_method_ws,
            payment_status_ws=payment_status_ws,
            units_ws=units_ws,
            vat_rates_ws=vat_rates_ws,
            actions_ws=actions_ws
        )

        DATA_SHEETS = {
            cfg.INVOICE_METADATA_SHEET_NAME,
            cfg.INVOICE_ITEMS_SHEET_NAME,
        }

        for ws in wb.worksheets:
            self.style_header(ws)
            self.autosize_columns(ws)

            if ws.title in DATA_SHEETS:
                self.freeze_header(ws)
                self.apply_autofilter(ws)
                self.zebra_rows(ws)

        # opcjonalnie: usuń domyślny pusty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)

    def _write_invoices(self, wb: Workbook, invoices) -> Worksheet:
        ws = wb.create_sheet(cfg.INVOICE_METADATA_SHEET_NAME)

        headers = [
            "action",  # APPLY | MODIFY | DELETE
            "invoice_number",  # nowy numer (lub pusty → generator)
            "old_invoice_number",  # 🔒 UKRYTA – tylko dla MODIFY
            "invoice_date",
            "selling_date",
            "buyer_NIP",
            "seller_NIP",
            "payment_method",
            "payment_status",
            "due_date",
            "timestamp",
        ]
        ws.append(headers)

        for i in invoices:
            ws.append([
                str(InvoiceCommand.APPLY.value),
                i.invoice_number,
                i.invoice_number,  # old_invoice_number
                i.invoice_date,
                i.selling_date,
                str(i.buyer_tax_number) if i.buyer_tax_number else None,
                str(i.seller_tax_number) if i.seller_tax_number else None,
                i.payment_method.value if i.payment_method else None,
                i.payment_status.value if i.payment_status else None,
                i.due_date,
                i.timestamp,
            ])
        ws.column_dimensions["C"].hidden = True
        return ws

    def _write_invoice_lines(self, wb: Workbook, lines) -> Worksheet:
        ws = wb.create_sheet(cfg.INVOICE_ITEMS_SHEET_NAME)

        headers = [
            "id",
            "invoice_number",
            "item_name",
            "description",
            "quantity",
            "unit",
            "net",
            "vat_rate",
            "tax_treatment",
            "contract_code",
            "cost_node_code",
            "cost_type_code",
        ]
        ws.append(headers)

        contracts = {
            c.id: c.code
            for c in self._contract_repository.list()
        }

        cost_nodes = {
            n.id: n.code
            for n in self._cost_node_repository.list_nodes()
        }

        cost_types = {
            t.id: t.code
            for t in self._cost_type_repository.list()
        }

        for l in lines:
            ws.append([
                str(l.id),
                str(l.invoice_number) if l.invoice_number else None,
                l.item_name,
                l.description,
                l.quantity,
                l.unit.value,
                l.net,
                l.vat_rate.name if isinstance(l.vat_rate, Enum) else l.vat_rate,
                l.tax_treatment.value,
                contracts.get(l.contract_id),
                cost_nodes.get(l.cost_node_id),
                cost_types.get(l.cost_type_id),
            ])
        ws.column_dimensions["A"].hidden = True
        return ws

    def _write_buyers(self, wb: Workbook, buyers) -> Worksheet:
        ws = wb.create_sheet(cfg.DICTS_BUYERS)
        ws.append(["tax_number","name","id" ])

        for c in buyers:
            ws.append([c.tax_number,c.name,str(c.id)  ])
        ws.column_dimensions["C"].hidden = True
        return ws

    def _write_sellers(self, wb: Workbook, sellers) -> Worksheet:
        ws = wb.create_sheet(cfg.DICTS_SELLERS)
        ws.append(["tax_number","name","id" ])

        for c in sellers:
            ws.append([c.tax_number,c.name,str(c.id)  ])
        ws.column_dimensions["C"].hidden = True
        return ws

    def _write_contracts(self, wb: Workbook, contracts) -> Worksheet:
        ws = wb.create_sheet(cfg.DICTS_CONTRACTS)
        ws.append(["code",  "name","id"])

        for c in contracts:
            ws.append([c.code, c.name ,str(c.id)])
        ws.column_dimensions["C"].hidden = True
        return ws

    def _write_cost_nodes(self, wb: Workbook, cost_nodes) -> Worksheet:
        ws = wb.create_sheet(cfg.DICTS_COST_NODES)
        ws.append([
            "code",
            "name",
            "budget",
            "id",
            "contract_id",
            "parent_id"
        ])

        for n in cost_nodes:
            ws.append([
                n.code,
                n.name,
                n.budget,
                str(n.id),
                str(n.contract_id),
                str(n.parent_id) if n.parent_id else None
            ])
        ws.column_dimensions["D"].hidden = True  # id
        ws.column_dimensions["E"].hidden = True  # contract_id
        ws.column_dimensions["F"].hidden = True  # parent_id
        return ws

    def _write_cost_types(self, wb: Workbook, cost_types) -> Worksheet:
        ws = wb.create_sheet(cfg.DICTS_COST_TYPES)
        ws.append(["code", "name","id"])

        for ct in cost_types:
            ws.append([ ct.code, ct.name,str(ct.id)])

        ws.column_dimensions["C"].hidden = True
        return ws

    @staticmethod
    def _write_dictionary( wb: Workbook, dictionary: dict[str,str],dicts_sheet_name: str) -> Worksheet:
        ws = wb.create_sheet(dicts_sheet_name)
        ws.append(["label","code" ])

        for code,label in dictionary.items():
            ws.append([label,code])

        return ws

    def _apply_dropdowns(
        self,
        *,
        invoices_ws:Worksheet,
        lines_ws: Worksheet,
        buyers_ws: Worksheet,
        sellers_ws: Worksheet,
        contracts_ws: Worksheet,
        cost_nodes_ws: Worksheet,
        cost_types_ws: Worksheet,
        tax_treatments_ws: Worksheet,
        payment_method_ws: Worksheet,
        payment_status_ws: Worksheet,
        units_ws: Worksheet,
        vat_rates_ws: Worksheet,
        actions_ws: Worksheet,

    ) -> None:

        max_rows = 2000  # bezpieczny limit

        self._apply_one_dropdown(
            max_rows,
            actions_ws,
            cfg.DICTS_ACTIONS,
            invoices_ws,
            "A"
        )

        self._apply_one_dropdown(
            max_rows,
            buyers_ws,
            cfg.DICTS_BUYERS,
            invoices_ws,
            "F"
        )

        self._apply_one_dropdown(
            max_rows,
            sellers_ws,
            cfg.DICTS_SELLERS,
            invoices_ws,
            "G"
        )
        self._apply_one_dropdown(
            max_rows,
            payment_method_ws,
            cfg.DICTS_PAYMENT_METHODS,
            invoices_ws,
            "H"
        )
        self._apply_one_dropdown(
            max_rows,
            payment_status_ws,
            cfg.DICTS_PAYMENT_STATUS,
            invoices_ws,
            "I"
        )

        self._apply_one_dropdown(
            max_rows,
            units_ws,
            cfg.DICTS_UNITS,
            lines_ws,
            "F"
        )


        self._apply_one_dropdown(
            max_rows,
            vat_rates_ws,
            cfg.DICTS_VAT_RATES,
            lines_ws,
            "H",
            "B"
        )

        self._apply_one_dropdown(
            max_rows,
            tax_treatments_ws,
            cfg.DICTS_TAX_TREATMENTS,
            lines_ws,
            "I"
        )

        self._apply_one_dropdown(
            max_rows,
            contracts_ws,
            cfg.DICTS_CONTRACTS,
            lines_ws,
            "J"
        )

        self._apply_one_dropdown(
            max_rows,
            cost_nodes_ws,
            cfg.DICTS_COST_NODES,
            lines_ws,
            "K"
        )

        self._apply_one_dropdown(
            max_rows,
            cost_types_ws,
            cfg.DICTS_COST_TYPES,
            lines_ws,
            "L"
        )


    @staticmethod
    def _apply_one_dropdown(max_rows: int,
                            dict_ws: Worksheet,
                            dict_ws_name: str,
                            source_ws: Worksheet,
                            target_column: str,
                            source_column: str = "A"
                            ) -> None:
        data_range = f"{dict_ws_name}!${source_column}$2:${source_column}${dict_ws.max_row}"
        dv = DataValidation(
            type="list",
            formula1=f"={data_range}",
            allow_blank=True,
        )
        source_ws.add_data_validation(dv)
        dv.add(f"{target_column}2:{target_column}{max_rows}")

    @staticmethod
    def autosize_columns(ws: Worksheet, max_width: int = 50) -> None:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 4, max_width +1)


    @staticmethod
    def style_header(
            ws: Worksheet,
            header_row: int = 1,
            bg_color: str = "1F4E79",  # ciemny niebieski
            font_color: str = "FFFFFF",
    ) -> None:
        header_font = Font(bold=True, color=font_color)
        header_fill = PatternFill(
            fill_type="solid",
            start_color=bg_color,
            end_color=bg_color,
        )
        header_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


    @staticmethod
    def zebra_rows(
            ws: Worksheet,
            start_row: int = 2,  # od pierwszego wiersza danych
            bg_color: str = "EAF2FB"  # bardzo jasny niebieski
    ) -> None:

        if ws.max_row < start_row:
            return

        fill = PatternFill(
            fill_type="solid",
            start_color=bg_color,
            end_color=bg_color,
        )

        rule = FormulaRule(
            formula=[f"MOD(ROW(),2)=0"],
            fill=fill,
        )

        end_col = get_column_letter(ws.max_column)

        ws.conditional_formatting.add(
            f"A{start_row}:{end_col}{ws.max_row}",
            rule,
        )

    @staticmethod
    def freeze_header(ws: Worksheet) -> None:
        ws.freeze_panes = "A2"

    @staticmethod
    def apply_autofilter(ws: Worksheet) -> None:
        if ws.max_row < 2:
            return  # brak danych

        end_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{end_col}{ws.max_row}"
