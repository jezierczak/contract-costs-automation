from pathlib import Path
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font


class ExcelReportRenderer:
    NUMERIC_COLUMNS = {
        "cost_node_budget",
        "net_amount",
        "vat_amount",
        "gross_amount",
        "non_tax_amount",
        "quantity",
        "total",
        "earned"
    }

    @staticmethod
    def render(
        df: pd.DataFrame,
        *,
        output_path: Path,
        sheet_name: str = "report",
    ) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.book[sheet_name]

            # mapowanie: nazwa kolumny -> litera excela
            col_letters = {
                col_name: ExcelReportRenderer._excel_col_letter(idx + 1)
                for idx, col_name in enumerate(df.columns)
            }

            for col_name in ExcelReportRenderer.NUMERIC_COLUMNS:
                if col_name not in col_letters:
                    continue

                col_letter = col_letters[col_name]

                for cell in ws[col_letter][1:]:  # pomijamy nagłówek
                    if cell.value is not None:
                        cell.number_format = "#,##0.00"  # 🇵🇱 Excel format

            # earned < 0 → czerwone tło
            if "earned" in col_letters:
                col_letter = col_letters["earned"]
                start_row = 2  # dane (bez nagłówka)
                end_row = ws.max_row

                red_fill = PatternFill(
                    start_color="FFC7CE",
                    end_color="FFC7CE",
                    fill_type="solid",
                )
                red_font = Font(color="9C0006")

                ws.conditional_formatting.add(
                    f"{col_letter}{start_row}:{col_letter}{end_row}",
                    CellIsRule(
                        operator="lessThan",
                        formula=["0"],
                        fill=red_fill,
                        font=red_font,
                    ),
                )
            # format SUMA (pogrubienie + górna linia)
            bold_font = Font(bold=True)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                first_cell = row[0]
                if first_cell.value == "SUMA":
                    for cell in row:
                        cell.font = bold_font

    @staticmethod
    def _excel_col_letter(n: int) -> str:
        """1 -> A, 2 -> B, ..., 27 -> AA"""
        result = ""
        while n:
            n, rem = divmod(n - 1, 26)
            result = chr(65 + rem) + result
        return result
