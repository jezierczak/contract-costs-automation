from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook

from contract_costs.infrastructure.excel.invoice_excel_context import InvoiceExcelContext, EXCEL_SPECS
from contract_costs.services.invoices.actions.dto.invoice_action_command import InvoiceActionCommand, InvoiceSelector, \
    InvoiceAction, invoice_action_from_excel


class InvoiceActionExcelLoader:


    @staticmethod
    def load(
            path: Path,
            *,
            context: InvoiceExcelContext,
    ) -> list[InvoiceActionCommand]:

        spec = EXCEL_SPECS[context]

        wb = load_workbook(path)
        ws = wb.active

        grouped: dict[str, list[InvoiceSelector]] = {}

        for row in ws.iter_rows(min_row=2):

            raw_action = row[spec.action_column].value
            invoice_id = row[spec.invoice_id_column].value

            if not raw_action or not invoice_id:
                continue

            raw_action = str(raw_action).strip().lower()
            if raw_action not in spec.allowed_actions:
                continue

                # raise ValueError(
                #     f"Action '{raw_action}' not allowed in {context.value} excel"
                # )

            grouped.setdefault(raw_action, []).append(
                InvoiceSelector(invoice_id=UUID(str(invoice_id)))
            )

        return [
            InvoiceActionCommand(
                action=invoice_action_from_excel(
                    context=context,
                    raw=raw_action,
                ),
                selectors=selectors,
                payload=None,
            )
            for raw_action, selectors in grouped.items()
        ]
