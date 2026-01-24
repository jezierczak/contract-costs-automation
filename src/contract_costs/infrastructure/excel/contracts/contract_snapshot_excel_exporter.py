from decimal import Decimal
from pathlib import Path
from uuid import UUID
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from contract_costs.infrastructure.excel.excel_common_methods import ExcelCommonMethods
from contract_costs.services.snapshots.dto.contract_snapshot_dto import (
    ContractSnapshotDTO,
    ContractNodeSnapshotDTO,
)

# =========================================================
# STYLES
# =========================================================

FONT_ROOT = Font(bold=True, color="1F4E79")
FONT_GROUP = Font(bold=True)
FONT_LEAF = Font(bold=False)

HEADER_FONT = Font(bold=True)
CENTER = Alignment(vertical="center")

# =========================================================
# TREE INDEX
# =========================================================

class SnapshotTreeIndex:
    def __init__(self, nodes: list[ContractNodeSnapshotDTO]):
        self.nodes = {n.node_id: n for n in nodes}
        self.children: dict[UUID | None, list[ContractNodeSnapshotDTO]] = defaultdict(list)

        for n in nodes:
            self.children[n.parent_id].append(n)

        for lst in self.children.values():
            lst.sort(key=lambda x: x.code)

    def roots(self) -> list[ContractNodeSnapshotDTO]:
        return self.children.get(None, [])

    def children_of(self, parent_id: UUID) -> list[ContractNodeSnapshotDTO]:
        return self.children.get(parent_id, [])

# =========================================================
# EXPORTER
# =========================================================

class ContractSnapshotExcelExporter:

    SHEET_NAME = "SNAPSHOT"

    def export(
        self,
        *,
        snapshot: ContractSnapshotDTO,
        output_path: Path,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME

        self._write_metadata(ws, snapshot)
        self._write_header(ws)

        tree = SnapshotTreeIndex(snapshot.nodes)

        start_row = ws.max_row + 1
        self._write_nodes(
            ws=ws,
            tree=tree,
            parent_id=None,
            prefix="",
        )

        # self._autosize(ws)
        ExcelCommonMethods.style_header(ws,5)
        ExcelCommonMethods.zebra_rows(ws,6)
        ExcelCommonMethods.freeze_header(ws,6)
        ExcelCommonMethods.autosize_columns(ws)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    # =========================================================
    # METADATA
    # =========================================================
    @staticmethod
    def _write_metadata( ws, snapshot: ContractSnapshotDTO) -> None:
        ws["A1"] = "Snapshot ID"
        ws["B1"] = str(snapshot.snapshot_id)

        ws["A2"] = "Snapshot Date"
        ws["B2"] = snapshot.snapshot_date

        ws["A3"] = "Contract"
        ws["B3"] = snapshot.contract_code

        for r in range(1, 4):
            ws[f"A{r}"].font = HEADER_FONT

        ws.append([])

    # =========================================================
    # HEADER
    # =========================================================
    @staticmethod
    def _write_header(ws) -> None:
        headers = [
            "CODE",
            "BUDGET",
            "PROG",
            "DONE",
            "NET",
            "NON-DEDUCT",
            "SPEND",
            "RESULT",
            "REV",
            "NAME",
        ]

        ws.append(headers)

        row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    # =========================================================
    # TREE RENDERING (DFS)
    # =========================================================

    def _write_nodes(
        self,
        *,
        ws,
        tree: SnapshotTreeIndex,
        parent_id: UUID | None,
        prefix: str,
    ) -> None:
        children = tree.children_of(parent_id)

        for idx, node in enumerate(children):
            last = idx == len(children) - 1
            connector = "└── " if last else "├── "
            label = f"{prefix}{connector}{node.code}"

            done = node.planned_budget * node.progress
            spend = node.net + node.non_deductible
            result = done - spend

            ws.append(
                [
                    label,
                    node.planned_budget,
                    node.progress,
                    done,
                    node.net,
                    node.non_deductible,
                    spend,
                    result,
                    node.revenue,
                    node.name,
                ]
            )

            row = ws.max_row
            self._style_row(ws, row, node, tree)
            self._format_row(ws, row)

            extension = "    " if last else "│   "
            self._write_nodes(
                ws=ws,
                tree=tree,
                parent_id=node.node_id,
                prefix=prefix + extension,
            )

    # =========================================================
    # STYLES
    # =========================================================

    @staticmethod
    def _style_row(

        ws,
        row: int,
        node: ContractNodeSnapshotDTO,
        tree: SnapshotTreeIndex,
    ) -> None:
        is_root = node.parent_id is None
        is_leaf = not tree.children_of(node.node_id)

        if is_root:
            font = FONT_ROOT
        elif is_leaf:
            font = FONT_LEAF
        else:
            font = FONT_GROUP

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = font

    @staticmethod
    def _format_row( ws, row: int) -> None:
        ws.cell(row=row, column=2).number_format = "#,##0.00"   # BUDGET
        ws.cell(row=row, column=3).number_format = "0.0%"      # PROG
        ws.cell(row=row, column=4).number_format = "#,##0.00"  # DONE
        ws.cell(row=row, column=5).number_format = "#,##0.00"  # NET
        ws.cell(row=row, column=6).number_format = "#,##0.00"  # NON-DEDUCT
        ws.cell(row=row, column=7).number_format = "#,##0.00"  # SPEND
        ws.cell(row=row, column=8).number_format = "#,##0.00"  # RESULT
        ws.cell(row=row, column=9).number_format = "#,##0.00"  # REV

    # =========================================================
    # HELPERS
    # =========================================================
    @staticmethod
    def _autosize( ws) -> None:
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["J"].width = 60

        for col in "BCDEFGHI":
            ws.column_dimensions[col].width = 14
