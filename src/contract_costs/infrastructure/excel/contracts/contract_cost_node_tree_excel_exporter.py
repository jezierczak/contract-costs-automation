
from pathlib import Path

from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


from contract_costs.infrastructure.excel.excel_common_methods import ExcelCommonMethods
from contract_costs.model.contract import Contract
from contract_costs.model.contract_node import ContractNode
from contract_costs.services.contracts.prepare.contract_node_tree_index import ContractNodeTreeIndex
from contract_costs.services.contracts.prepare.mappers.contract_node_prepare_mapper import ContractNodePrepareMapper

FONT_ROOT = Font(bold=True, color="1F4E79")
FONT_GROUP = Font(bold=True)
FONT_LEAF = Font(bold=False)
FONT_INACTIVE = Font(color="9E9E9E", italic=True)

class ContractTreeExcelExporter:
    """
    READ-ONLY Excel exporter.
    Used only for visual analysis of contract + cost node tree.
    NOT used for apply.
    """

    CONTRACT_SHEET = "CONTRACT"
    COST_NODES_SHEET = "COST_NODES"

    def export(
        self,
        *,
        contract: Contract,
        cost_nodes: list[ContractNode],
        output_path: Path,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Workbook has no active worksheet to remove")
        # remove default sheet
        wb.remove(ws)

        self._export_contract_sheet(wb, contract)
        self._export_cost_nodes_sheet(wb, cost_nodes)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    # =========================================================
    # CONTRACT SHEET
    # =========================================================

    def _export_contract_sheet(self, wb: Workbook, contract: Contract) -> None:
        ws = wb.create_sheet(self.CONTRACT_SHEET)

        header_font = Font(bold=True)

        rows = [
            ("Contract Code", contract.code),
            ("Contract Name", contract.name),
            ("Status", contract.status.value),
            ("Owner", contract.owner.name if contract.owner else ""),
            ("Client", contract.client.name if contract.client else ""),
            ("Start Date", contract.start_date),
            ("End Date", contract.end_date),
            ("Budget Declared", contract.budget),
        ]

        for row_idx, (label, value) in enumerate(rows, start=1):
            ws[f"A{row_idx}"] = label
            ws[f"A{row_idx}"].font = header_font
            ws[f"B{row_idx}"] = value

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40
        ExcelCommonMethods.autosize_columns(ws)
    # =========================================================
    # COST NODES SHEET
    # =========================================================

    def _export_cost_nodes_sheet(
        self,
        wb: Workbook,
        cost_nodes: list[ContractNode],
    ) -> None:
        ws = wb.create_sheet(self.COST_NODES_SHEET)

        headers = [
            "Node",
            "Code",
            "Name",
            "Quantity",
            "Unit",
            "Budget (own)",
            "Budget (total)",
            "Active",
        ]

        ws.append(headers)

        header_font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        tree = ContractNodeTreeIndex(cost_nodes)

        self._write_node_rows(
            ws=ws,
            tree=tree,
            parent_id=None,
            prefix="",
            is_last=True,
        )

        ExcelCommonMethods.style_header(ws)
        ExcelCommonMethods.zebra_rows(ws)
        ExcelCommonMethods.freeze_header(ws)
        ExcelCommonMethods.autosize_columns(ws)

    # =========================================================
    # TREE RENDERING
    # =========================================================

    def _write_node_rows(
            self,
            *,
            ws,
            tree: ContractNodeTreeIndex,
            parent_id: UUID | None,
            prefix: str,
            is_last: bool,
    ) -> None:
        if parent_id is None:
            children = tree.roots()
        else:
            children = tree.children_of(parent_id)

        nodes_by_parent = tree.children_by_parent  # lokalny wyjątek

        for idx, node in enumerate(children):
            last = idx == len(children) - 1

            connector = "└── " if last else "├── "
            node_label = f"{prefix}{connector}{node.code}"

            total_budget = ContractNode.calculate_budget_from_leaves(
                node.id,
                nodes_by_parent,
            )

            ws.append(
                [
                    node_label,
                    node.code,
                    node.name,
                    node.quantity,
                    node.unit.value if node.unit else None,
                    node.budget,
                    total_budget,
                    "YES" if node.is_active else "NO",
                ]
            )
            row_idx = ws.max_row

            active_children = [
                c for c in tree.children_of(node.id)
                if c.is_active
            ]

            is_leaf = not active_children
            is_root = node.parent_id is None

            if not node.is_active:
                font = FONT_INACTIVE
            elif is_root:
                font = FONT_ROOT
            elif not is_leaf:
                font = FONT_GROUP
            else:
                font = FONT_LEAF

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).font = font

            extension = "    " if last else "│   "

            self._write_node_rows(
                ws=ws,
                tree=tree,
                parent_id=node.id,
                prefix=prefix + extension,
                is_last=last,
            )

    # =========================================================
    # HELPERS
    # =========================================================

    # @staticmethod
    # def _build_nodes_by_parent(
    #     nodes: Iterable[CostNode],
    # ) -> dict[UUID | None, list[CostNode]]:
    #     tree: dict[UUID | None, list[CostNode]] = defaultdict(list)
    #
    #     for node in nodes:
    #         tree[node.parent_id].append(node)
    #
    #     for children in tree.values():
    #         children.sort(key=lambda n: n.code)
    #
    #     return tree

