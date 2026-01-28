import re

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


class ExcelCommonMethods:

    @staticmethod
    def apply_one_dropdown(max_rows: int,
                           dict_ws: Worksheet,
                           dict_ws_name: str,
                           source_ws: Worksheet,
                           target_column: str,
                           source_column: str = "A"
                           ) -> None:
        data_range = f"{dict_ws_name}!${source_column}$2:${source_column}${dict_ws.max_row}"
        dv = DataValidation(
            type="list",
            formula1=f"={data_range}",
            allow_blank=True,
        )
        source_ws.add_data_validation(dv)
        dv.add(f"{target_column}2:{target_column}{max_rows}")

    @staticmethod
    def apply_formula_dropdown(
            *,
            max_rows: int,
            source_ws: Worksheet,
            target_column: str,
            formula: str,
    ) -> None:
        """
        Dropdown oparty o formułę (np. =INDIRECT($J2))
        """
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
        )
        source_ws.add_data_validation(dv)
        dv.add(f"{target_column}2:{target_column}{max_rows}")

    @staticmethod
    def autosize_columns(ws: Worksheet, max_width: int = 50) -> None:
        for idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(idx)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 4, max_width + 1)

    @staticmethod
    def style_header(
            ws: Worksheet,
            header_row: int = 1,
            bg_color: str = "1F4E79",  # ciemny niebieski
            font_color: str = "FFFFFF",
    ) -> None:
        header_font = Font(bold=True, color=font_color)
        header_fill = PatternFill(
            fill_type="solid",
            start_color=bg_color,
            end_color=bg_color,
        )
        header_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )

        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    @staticmethod
    def zebra_rows(
            ws: Worksheet,
            start_row: int = 2,  # od pierwszego wiersza danych
            bg_color: str = "EAF2FB"  # bardzo jasny niebieski
    ) -> None:

        if ws.max_row < start_row:
            return

        fill = PatternFill(
            fill_type="solid",
            start_color=bg_color,
            end_color=bg_color,
        )

        rule = FormulaRule(
            formula=[f"MOD(ROW(),2)=0"],
            fill=fill,
        )

        end_col = get_column_letter(ws.max_column)

        ws.conditional_formatting.add(
            f"A{start_row}:{end_col}{ws.max_row}",
            rule,
        )

    @staticmethod
    def freeze_header(ws: Worksheet, row: int = 2) -> None:
        ws.freeze_panes = f"A{row}"


    @staticmethod
    def apply_autofilter(ws: Worksheet) -> None:
        if ws.max_row < 2:
            return  # brak danych

        end_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{end_col}{ws.max_row}"

    @staticmethod
    def safe_named_range(name: str) -> str:
        """
        Excel named range rules:
        - must start with letter or underscore
        - cannot contain spaces
        - cannot look like cell reference
        """
        safe = re.sub(r"[^A-Za-z0-9_]", "_", str(name))

        if safe[0].isdigit():
            safe = f"_{safe}"

        return safe

    @staticmethod
    def apply_named_dropdown(
            *,
            ws,
            column: str,
            from_row: int,
            to_row: int,
            formula: str,
    ) -> None:
        """
        Apply dropdown validation to a column range.

        formula examples:
        - "=MY_NAMED_RANGE"
        - "=INDIRECT($B2)"
        """
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
        )

        ws.add_data_validation(dv)

        dv.add(f"{column}{from_row}:{column}{to_row}")