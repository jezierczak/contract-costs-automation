from pathlib import Path
from typing import Any, TypeVar, Generic
from openpyxl import load_workbook

from contract_costs.infrastructure.excel.excel_column import (
    ExcelColumn,
    ExcelColumnType,
)

T = TypeVar("T")


class ExcelLoader(Generic[T]):

    @staticmethod
    def load(
        *,
        input_path: Path,
        columns: list[ExcelColumn[T]],
        sheet_name: str | None = None,
        start_row: int = 2,
    ) -> list[dict[str, Any]]:

        wb = load_workbook(input_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        if ws is None:
            raise ValueError(
                f"No active worksheet found in Excel file: {input_path}"
            )

        column_map: dict[int, ExcelColumn[T]] = {
            idx + 1: col
            for idx, col in enumerate(columns)
        }

        rows: list[dict[str, Any]] = []

        for row_idx in range(start_row, ws.max_row + 1):
            row_data: dict[str, Any] = {}
            empty_row = True

            for col_idx, col in column_map.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value

                # normalize empty
                if value not in (None, ""):
                    empty_row = False

                value = ExcelLoader._normalize_cell(value)

                # =====================
                # CHECKBOX
                # =====================
                if col.column_type == ExcelColumnType.CHECKBOX:
                    value = ExcelLoader._parse_checkbox(
                        value, col.header
                    )

                row_data[col.header] = value

            if not empty_row:
                rows.append(row_data)

        return rows

    # =====================
    # HELPERS
    # =====================

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if value in ("", None):
            return None
        return value

    @staticmethod
    def _parse_checkbox(
        value: Any,
        header: str,
    ) -> bool:
        if value is None or value == "":
            return False

        if isinstance(value, bool):
            return value

        value_str = str(value).strip().upper()

        if value_str == "YES":
            return True

        if value_str == "NO":
            return False

        raise ValueError(
            f"Invalid checkbox value '{value}' "
            f"for column '{header}'. Expected YES or NO."
        )
