
from pathlib import Path
from typing import Generic, TypeVar

from openpyxl import Workbook
from openpyxl.styles import Protection, Alignment
from openpyxl.utils import get_column_letter

from contract_costs.infrastructure.excel.checkbox_options import CheckBoxOptions
from contract_costs.infrastructure.excel.excel_column import (
    ExcelColumn,
    ExcelColumnType,
)
from contract_costs.infrastructure.excel.excel_common_methods import (
    ExcelCommonMethods,
)
import contract_costs.config as cfg

T = TypeVar("T")


class BaseExcelExporter(Generic[T]):
    """
    Excel exporter supporting:
    - single-sheet export (static API)
    - multi-sheet export (instance API)

    Workbook lifecycle:
    - created once per exporter instance
    - multiple sheets can be added
    - saved explicitly via save()
    """

    # =========================================================
    # INIT
    # =========================================================

    def __init__(self) -> None:
        self._wb = Workbook()
        self._initialized = False

    # =========================================================
    # PUBLIC API (INSTANCE – MULTI SHEET)
    # =========================================================

    def add_sheet(
        self,
        *,
        items: list[T],
        columns: list[ExcelColumn[T]],
        sheet_name: str,
        style: bool = True,
    ) -> None:
        """
        Add a sheet to the workbook.

        Does NOT save the file.
        """
        # first sheet → reuse active
        if not self._initialized:
            ws = self._wb.active
            self._initialized = True
        else:
            ws = self._wb.create_sheet()

        if ws is None:
            raise RuntimeError("Workbook has no active worksheet")
        ws.title = sheet_name

        # ======================
        # HEADER
        # ======================
        ws.append([col.header for col in columns])

        # ======================
        # ROWS
        # ======================
        for item in items:
            row = []
            for col in columns:
                try:
                    value = col.getter(item)
                except Exception as e:
                    raise RuntimeError(
                        f"Error while exporting column '{col.header}'"
                    ) from e

                match col.column_type:

                    case ExcelColumnType.CHECKBOX:
                        value = CheckBoxOptions.YES.value if value else CheckBoxOptions.NO.value

                    case ExcelColumnType.LINK:
                        if value:
                            abs_path = (cfg.WORK_DIR / Path(value)).resolve().as_posix()
                            value = f'=HYPERLINK("file:///{abs_path}", "📄 Otwórz")'
                        else:
                            value = None

                    case ExcelColumnType.FOLDER:
                        if value:
                            folder = (cfg.WORK_DIR / Path(value).parent).resolve().as_posix()
                            value = f'=HYPERLINK("file:///{folder}", "📂 Folder")'
                        else:
                            value = None

                    case _:
                        pass  # DISPLAY, HIDDEN, DROPDOWN itd.

                row.append(value)

            ws.append(row)

        # ======================
        # COLUMN METADATA
        # ======================
        for idx, col in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)

            # --- HIDDEN ---
            if col.column_type == ExcelColumnType.HIDDEN:
                ws.column_dimensions[col_letter].hidden = True

            # --- PROTECTION ---
            for (cell,) in ws.iter_rows(
                min_col=idx,
                max_col=idx,
                min_row=2,
                max_row=ws.max_row,
            ):
                cell.protection = Protection(
                    locked=not col.editable
                )

        # ======================
        # STYLE
        # ======================
        if style:
            ExcelCommonMethods.style_header(ws)
            ExcelCommonMethods.autosize_columns(ws)
            ExcelCommonMethods.freeze_header(ws)
            ExcelCommonMethods.zebra_rows(ws)

        # align first column (common UX)
        for cell in ws["A"]:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # ======================
        # DROPDOWNS
        # ======================
        for idx, col in enumerate(columns, start=1):
            if col.column_type not in (
                    ExcelColumnType.DROPDOWN,
                    ExcelColumnType.CHECKBOX,
            ):
                continue

                # -----------------------------------
                # RESOLVE OPTIONS
                # -----------------------------------
            if col.column_type == ExcelColumnType.CHECKBOX:
                options = [
                    CheckBoxOptions.YES.value,
                    CheckBoxOptions.NO.value,
                ]
            else:
                if not col.options:
                    raise RuntimeError(
                        f"Dropdown column '{col.header}' has no options"
                    )
                options = col.options

            col_letter = get_column_letter(idx)

            dict_ws = self._wb.create_sheet(f"_dict_{sheet_name}_{col.header}")
            dict_ws.sheet_state = "hidden"

            dict_ws["A1"] = col.header
            for row_idx, opt in enumerate(options, start=2):
                dict_ws[f"A{row_idx}"] = opt

            ExcelCommonMethods.apply_one_dropdown(
                max_rows=max(ws.max_row, 2),
                dict_ws=dict_ws,
                dict_ws_name=dict_ws.title,
                source_ws=ws,
                target_column=col_letter,
            )

    def save(self, output_path: Path) -> None:
        """
        Save workbook to disk.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(output_path)

    # =========================================================
    # BACKWARD COMPATIBLE API (STATIC – SINGLE SHEET)
    # =========================================================

    @staticmethod
    def export(
        *,
        items: list[T],
        columns: list[ExcelColumn[T]],
        output_path: Path,
        sheet_name: str = "data",
    ) -> None:
        """
        Backward compatible single-sheet export.
        """
        exporter = BaseExcelExporter[T]()
        exporter.add_sheet(
            items=items,
            columns=columns,
            sheet_name=sheet_name,
        )
        exporter.save(output_path)

    # =========================================================
    # OPTIONAL SUGAR
    # =========================================================

    def export_many(
        self,
        *,
        sheets: list[tuple[str, list[T], list[ExcelColumn[T]]]],
        output_path: Path,
    ) -> None:
        """
        Convenience method for exporting many sheets at once.
        """
        for sheet_name, items, columns in sheets:
            self.add_sheet(
                sheet_name=sheet_name,
                items=items,
                columns=columns,
            )

        self.save(output_path)
