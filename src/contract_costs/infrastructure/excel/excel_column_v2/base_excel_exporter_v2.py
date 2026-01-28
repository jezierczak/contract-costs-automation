from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Protection, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.cell import WriteOnlyCell

from contract_costs.infrastructure.excel.checkbox_options import CheckBoxOptions
from contract_costs.infrastructure.excel.excel_column_v2.excel_column import ExcelColumn
from contract_costs.infrastructure.excel.excel_column_v2.excel_column_type import   ExcelColumnType

from contract_costs.infrastructure.excel.excel_common_methods import (
    ExcelCommonMethods,
)
import contract_costs.config as cfg


FONT_ROOT = Font(bold=True, color="1F4E79")
FONT_GROUP = Font(bold=True)
FONT_LEAF = Font(bold=False)
FONT_INACTIVE = Font(color="9E9E9E", italic=True)

class BaseExcelExporterV2[T]:
    """
    Excel exporter supporting:
    - single-sheet export (static API)
    - multi-sheet export (instance API)

    Workbook lifecycle:
    - created once per exporter instance
    - multiple sheets can be added
    - saved explicitly via save()
    """

    # =========================================================
    # INIT
    # =========================================================

    def __init__(self) -> None:
        self._wb = Workbook()
        self._initialized = False
        # registry słowników
        self._dictionaries: dict[str, Worksheet] = {}

    # =========================================================
    # PUBLIC API (INSTANCE – MULTI SHEET)
    # =========================================================

    def add_sheet(
        self,
        *,
        items: list[T],
        columns: list[ExcelColumn[T]],
        sheet_name: str,
        header: dict[str, list[str]] | None = None,
        style: bool = True,
    ) -> None:
        """
        Add a sheet to the workbook.

        Does NOT save the file.
        """
        # first sheet → reuse active
        if not self._initialized:
            ws = self._wb.active
            self._initialized = True
        else:
            ws = self._wb.create_sheet()

        if ws is None:
            raise RuntimeError("Workbook has no active worksheet")
        ws.title = sheet_name

        # ======================
        # HEADER
        # ======================
        if header:
            header_rows = len(header)
            table_header_row = header_rows + 2  # 1 linia odstępu
        else:
            header_rows = 0
            table_header_row = 1

        data_start_row = table_header_row + 1

        # ======================
        # REPORT HEADER
        # ======================
        if header:
            row_idx = 1
            bold = Font(bold=True)

            for key, values in header.items():
                cell = ws.cell(row=row_idx, column=1, value=key)
                cell.font = bold

                for col_offset, val in enumerate(values, start=2):
                    ws.cell(row=row_idx, column=col_offset, value=val)

                row_idx += 1


        dropdown_columns = [
            col for col in columns
            if col.column_type == ExcelColumnType.DROPDOWN and col.dropdown
        ]


        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=table_header_row, column=col_idx, value=col.header)

        tree_col_idx: int | None = None
        tree_options = None

        tree_columns = [
            (idx, col) for idx, col in enumerate(columns)
            if col.column_type == ExcelColumnType.TREE
        ]

        if len(tree_columns) > 1:
            raise RuntimeError("Only one TREE column is allowed")

        if tree_columns:
            tree_col_idx, col = tree_columns[0]
            if col.tree is None:
                raise RuntimeError("TREE column requires TreeOptions")
            tree_options = col.tree
        else:
            tree_col_idx = None

        # ======================
        # ROWS
        # ======================
        if tree_col_idx is None:
            self._write_flat_rows(
                ws=ws,
                items=items,
                columns=columns,
                start_row=data_start_row,
            )
        else:
            self._write_tree_rows(
                ws=ws,
                items=items,
                columns=columns,
                tree_col_idx=tree_col_idx,
                tree=tree_options,
                start_row=data_start_row,
            )

        self._append_agg_row(
            ws=ws,
            columns=columns,
            data_start_row=data_start_row,
        )

        # ======================
        # COLUMN METADATA
        # ======================
        for idx, col in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)

            # --- HIDDEN ---
            if col.column_type == ExcelColumnType.HIDDEN:
                ws.column_dimensions[col_letter].hidden = True

            if col.column_type == ExcelColumnType.DISPLAY:
                for (cell,) in ws.iter_rows(
                        min_col=idx,
                        max_col=idx,
                        min_row=data_start_row,
                        max_row=ws.max_row,
                ):
                    if isinstance(cell.value, (int, float, Decimal)):
                        cell.number_format = '#,##0.00'

            if col.column_type == ExcelColumnType.PERCENT:
                for (cell,) in ws.iter_rows(
                        min_col=idx,
                        max_col=idx,
                        min_row=data_start_row,
                        max_row=ws.max_row,
                ):
                    cell.number_format = "0.0%"

            # --- PROTECTION ---
            for (cell,) in ws.iter_rows(
                min_col=idx,
                max_col=idx,
                min_row=data_start_row,
                max_row=ws.max_row,
            ):
                cell.protection = Protection(
                    locked=not col.editable
                )

        # ======================
        # STYLE
        # ======================
        if style:
            ExcelCommonMethods.style_header(ws, header_row=table_header_row)
            ExcelCommonMethods.autosize_columns(ws)
            ExcelCommonMethods.freeze_header(ws,row=data_start_row)
            ExcelCommonMethods.zebra_rows(ws, start_row=data_start_row)

        # align first column (common UX)
        for idx, col in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)

            if col.column_type == ExcelColumnType.TREE:
                for (cell,) in ws.iter_rows(
                        min_col=idx,
                        max_col=idx,
                        min_row=data_start_row,
                        max_row=ws.max_row,
                ):
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                    )

        # ======================
        # DROPDOWN DICTIONARIES
        # ======================
        named_ranges_by_dict: dict[str, dict[str, str]] = {}

        for col in dropdown_columns:
            if not col.dropdown:
                raise ValueError("Dropdown column requires DropdownOptions")
            dict_name = col.dropdown.dictionary
            if dict_name not in named_ranges_by_dict:
                if dict_name not in self._dictionaries:
                    raise RuntimeError(f"Dictionary '{dict_name}' not registered")
                dict_ws = self._dictionaries[dict_name]
                named_ranges_by_dict[dict_name] = (
                    self._create_named_ranges_from_dictionary(
                        ws=dict_ws,
                        value_column=col.dropdown.value_column,
                    )
                )
        # ======================
        # DROPDOWNS
        # ======================
        for idx, col in enumerate(columns, start=1):
            if col.column_type != ExcelColumnType.DROPDOWN:
                continue

            dropdown = col.dropdown
            if not dropdown:
                continue
            col_letter = get_column_letter(idx)

            from_row = data_start_row
            to_row = max(ws.max_row, data_start_row)

            if dropdown.depends_on is None:
                # statyczny dropdown
                ranges = named_ranges_by_dict[dropdown.dictionary]

                if len(ranges) != 1:
                    raise RuntimeError(
                        f"Static dropdown '{col.name}' requires exactly one KEY in dictionary "
                        f"'{dropdown.dictionary}'"
                    )

                named_range = next(iter(ranges.values()))

                ExcelCommonMethods.apply_named_dropdown(
                    ws=ws,
                    column=col_letter,
                    from_row=from_row,
                    to_row=to_row,
                    formula=f"={named_range}",
                )

            else:
                # dropdown relacyjny
                src_idx = next(
                    i + 1
                    for i, c in enumerate(columns)
                    if c.name == dropdown.depends_on
                )
                src_letter = get_column_letter(src_idx)

                ExcelCommonMethods.apply_named_dropdown(
                    ws=ws,
                    column=col_letter,
                    from_row=from_row,
                    to_row=to_row,
                    formula=f"=INDIRECT(${src_letter}{from_row})",
                )

        # --- ENABLE SHEET PROTECTION ---
        ws.protection.enable()

    def save(self, output_path: Path) -> None:
        """
        Save workbook to disk.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(output_path)

    # =========================================================
    # BACKWARD COMPATIBLE API (STATIC – SINGLE SHEET)
    # =========================================================

    @staticmethod
    def export(
        *,
        items: list[T],
        columns: list[ExcelColumn[T]],
        output_path: Path,
        sheet_name: str = "data",
    ) -> None:
        """
        Backward compatible single-sheet export.
        """
        exporter = BaseExcelExporterV2[T]()
        exporter.add_sheet(
            items=items,
            columns=columns,
            sheet_name=sheet_name,
        )
        exporter.save(output_path)

    # =========================================================
    # OPTIONAL SUGAR
    # =========================================================

    def export_many(
        self,
        *,
        sheets: list[tuple[str, list[T], list[ExcelColumn[T]]]],
        output_path: Path,
    ) -> None:
        """
        Convenience method for exporting many sheets at once.
        """
        for sheet_name, items, columns in sheets:
            self.add_sheet(
                sheet_name=sheet_name,
                items=items,
                columns=columns,
            )

        self.save(output_path)

    def register_dictionary(
            self,
            *,
            name: str,
            rows: list[dict[str, str]],
            hidden: bool = False,
    ) -> None:
        """
        rows example:
        [
            {"KEY": "C1", "VALUE": "Node A", "DESC": "Fundament"},
            {"KEY": "C1", "VALUE": "Node B", "DESC": "Ściany"},
        ]
        """
        ws = self._wb.create_sheet(f"_dict_{name}")
        ws.sheet_state = "hidden" if hidden else "visible"

        headers = list(rows[0].keys())
        ws.append(headers)

        for row in rows:
            ws.append([row[h] for h in headers])

        self._dictionaries[name] = ws

    def _create_named_ranges_from_dictionary(
            self,
            *,
            ws: Worksheet,
            key_column: str = "KEY",
            value_column: str = "VALUE",
    ) -> dict[str, str]:
        """
        Returns:
            key -> named_range_name
        """
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        key_col = headers[key_column]
        val_col = headers[value_column]

        ranges: dict[str, list[int]] = {}

        for row in range(2, ws.max_row + 1):
            key_raw = ws.cell(row=row, column=key_col).value
            if key_raw is None:
                continue  # albo raise, zależnie od reguł

            key = str(key_raw)
            ranges.setdefault(key, []).append(row)

        named_ranges: dict[str, str] = {}

        for key, rows in ranges.items():
            start = min(rows)
            end = max(rows)

            col_letter = get_column_letter(val_col)
            range_ref = f"{ws.title}!${col_letter}${start}:${col_letter}${end}"

            safe_name = ExcelCommonMethods.safe_named_range(f"{ws.title}_{key}")
            defined_name = DefinedName(
                name=safe_name,
                attr_text=range_ref,
            )
            self._wb.defined_names.add(defined_name)

            named_ranges[key] = safe_name

        return named_ranges

    def _write_flat_rows(
            self,
            *,
            ws,
            items: list[T],
            columns: list[ExcelColumn[T]],
            start_row: int,
    ) -> None:
        current_row = start_row

        for item in items:
            row = []
            for col in columns:
                value = col.getter(item)

                match col.column_type:
                    case ExcelColumnType.CHECKBOX:
                        value = CheckBoxOptions.YES.value if value else CheckBoxOptions.NO.value
                    case ExcelColumnType.LINK:
                        if value:
                            abs_path = (cfg.WORK_DIR / Path(value)).resolve().as_posix()
                            value = f'=HYPERLINK("file:///{abs_path}", "📄 Otwórz")'
                        else:
                            value = None
                    case ExcelColumnType.FOLDER:
                        if value:
                            folder = (cfg.WORK_DIR / Path(value).parent).resolve().as_posix()
                            value = f'=HYPERLINK("file:///{folder}", "📂 Folder")'
                        else:
                            value = None

                row.append(value)

            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

    @staticmethod
    def _write_tree_rows(

            *,
            ws,
            items: list[T],
            columns: list[ExcelColumn[T]],
            tree_col_idx: int,
            tree,
            start_row: int,
    ) -> None:
        # build children map
        children: dict[object | None, list[T]] = {}
        for item in items:
            pid = tree.parent_id(item)
            children.setdefault(pid, []).append(item)

        if tree.sort_key:
            for lst in children.values():
                lst.sort(key=tree.sort_key)

        current_row = start_row

        def walk(parent_id, prefix: str):
            nonlocal current_row
            nodes = children.get(parent_id, [])

            for idx, item in enumerate(nodes):
                last = idx == len(nodes) - 1
                connector = "└── " if last else "├── "
                extension = "    " if last else "│   "

                row = []
                for c_idx, col in enumerate(columns):
                    if c_idx == tree_col_idx:
                        if parent_id is None:
                            label = connector + str(col.getter(item))
                        else:
                            label = prefix + connector + str(col.getter(item))
                        row.append(label)
                    else:
                        row.append(col.getter(item))

                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=current_row, column=col_idx, value=value)

                # ==== STYL (ZACHOWUJE TWÓJ LOOK) ====
                active = tree.is_active(item) if tree.is_active else True
                has_children = bool(children.get(tree.id(item)))
                is_root = parent_id is None

                if not active:
                    font = FONT_INACTIVE
                elif is_root:
                    font = FONT_ROOT
                elif has_children:
                    font = FONT_GROUP
                else:
                    font = FONT_LEAF

                for column in range(1, ws.max_column + 1):
                    ws.cell(row=current_row, column=column).font = font
                # ===================================

                current_row += 1
                walk(tree.id(item), prefix + extension)

        walk(None, "")

    @staticmethod
    def _append_agg_row(
            *,
            ws: Worksheet,
            columns: list[ExcelColumn],
            data_start_row: int,
    ) -> None:
        if not any(col.agg for col in columns):
            return

        last_data_row = ws.max_row  # 🔒 ZAMRAŻAMY
        agg_row = last_data_row + 1

        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=agg_row, column=idx)

            if col.column_type == ExcelColumnType.TREE:
                cell.value = "SUM"
                cell.font = FONT_GROUP
                continue

            if not col.agg:
                continue

            col_letter = get_column_letter(idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
            cell.font = FONT_GROUP
